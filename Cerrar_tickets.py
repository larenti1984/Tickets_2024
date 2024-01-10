# Librerías
#from selenium import webdriver
#from selenium.webdriver.support.ui import WebDriverWait
#from selenium.webdriver.support import expected_conditions as EC
#from selenium.webdriver.common.by import By
#import time
#para listas
#from selenium.webdriver.support.ui import Select
#from openpyxl import Workbook
#from openpyxl import load_workbook
import openpyxl
import codigo
import datetime
import time

# Abrir el archivo de Excel
wb = openpyxl.load_workbook('tickets_creados.xlsx')

# Seleccionar la hoja de trabajo
sheet = wb['TICKETS']

# Inicializa las variables para recorrer las columnas
#col1 = 1
#col2 = 2
col3 = 3
col4 = 4
row = 2
ultima_row = sheet.max_row + 1


while True:
    # Obtener el valor de la celda
    #cell1_title = sheet.cell(row=row, column=col1).value
    cell3_nrotk = sheet.cell(row=row, column=col3).value

    # Si ambas celdas están vacías, detén el ciclo
    if not cell3_nrotk:
       break
    
     #cargar_driver()
    codigo.cargar_driver('https://wecare.neoris.net/')

    # click en My Activities ok
    codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/div/table[3]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/table[1]/tbody/tr[1]/td[3]")

    # input con umero de ticket
    codigo.d_ID_text(cell3_nrotk, 'rcTicket')
    
    # click en ticket buscado
    codigo.d_ID("tblProdServ")
    
    # click en Ticket Activities
    codigo.d_Xpatch("//html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div/div/div[1]/div[1]/ul/li[2]/a[2]/em/span/span")
    
     # click en Aceptar
    codigo.d_ID("btnAccept")
    #codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div/div/div[2]/div/div[3]/div/div/div/table/tbody/tr[2]/td[5]/input[1]")
    
    # input con "Cerrado"
    codigo.d_ID_text("Ticket Aceptado", 'txtComment')
    
     # click en Save
    #codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td/table/tbody/tr/td/input[1]")
    codigo.d_ID("btnOK")

    # click en Ticket activities 2 
    codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div/div/div[1]/div[1]/ul/li[2]/a[2]/em/span/span")

    # Hours
    #codigo.d_I_text("1", "txtTicketHours")
    codigo.d_SKcss_txt('1', 'input#txtTicketTime')


    # Ticket Comment
    #codigo.d_I_text("Horas", "txtTicketComments")
    codigo.d_SKcss_txt('1', 'input#txtTicketComments')

    # Click en Add Hours
    codigo.d_ID('btnAddTime')

    
    # click en Ok pop Alert okkkkkk
    time.sleep(10)
    codigo.ok_input()
   

    # Click en Finalice
    #codigo.d_css_click('input#btnFinalize')
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div/div/div[2]/div/div/div/div/div/table/tbody/tr[5]/td[3]/input')
                    

    
    # Ticket Type
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[3]/td/fieldset/table/tbody/tr[1]/td[2]/select/option[2]')
                     
    # Loading Channel
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[3]/td/fieldset/table/tbody/tr[2]/td[2]/select/option[5]')
    
    # Cause Type
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[3]/td/fieldset/table/tbody/tr[3]/td[2]/select/option[2]')
    
    # Description
    codigo.d_ID_text('.', 'txtComment')

      



    # click en Finalice
    #codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td/div/div/div[2]/div/div[2]/div/div/div/table/tbody/tr[5]/td[3]/input")
    
    # input con "Cerrado"
    #codigo.d_ID_text("Ticket Cerrado.-", 'txtComment')
    
     # click en Save
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/table/tbody/tr/td/input[1]')
    
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%d/%m/%Y %H:%M:%S")

    nueva_row = ultima_row + 1
    sheet.cell(row=row, column=col4).value = formatted_time
    wb.save("tickets_Cerrados_Paso3.xlsx")

    # Incrementa la fila
    row += 1
    
    # Cierra el archivo de Excel
    
    #wb.close('tickets_credos.xlsx')
    codigo.Cerrar_Driver()