import openpyxl
import codigo
import datetime

# Definir la fecha de caducidad del programa
fecha_caducidad = datetime.datetime(2024, 8, 30)

# Obtener la fecha actual
fecha_actual = datetime.datetime.now()

# Verificar si el programa ha caducado
if fecha_actual > fecha_caducidad:
    print("El programa ha caducado. Por favor, contacte al desarrollador para obtener una versión actualizada.")
    exit()

# Abrir el archivo de Excel
wb = openpyxl.load_workbook('crear_tickets.xlsx')

# Seleccionar la hoja de trabajo
sheet = wb['TICKETS']

# Inicializa las variables para recorrer las columnas
col1 = 1
col2 = 2
col3 = 3
# en esta columna iria el numero de categoria del ticket
col5 = 5
col6 = 6
col7 = 7
col8= 8
# la columna 9 es para que no se repita la carga de tickets
col9= 9



row = 2

while True:
    # Obtener el valor de la celda
    cell1_title = sheet.cell(row=row, column=col1).value
    cell2_desc = sheet.cell(row=row, column=col2).value
    cell_5cat = sheet.cell(row=row, column=col5).value
    cell_6cat = sheet.cell(row=row, column=col6).value
    cell_7cat = sheet.cell(row=row, column=col7).value
    cell_8cat = sheet.cell(row=row, column=col8).value
    cell_9cat = sheet.cell(row=row, column=col9).value

    # para no repetir los tickets
    if not cell_9cat == None:
      print ("estos ticket ya fueron realizados")
      exit ()



    # Si ambas celdas están vacías, detén el ciclo
    if not cell1_title or not cell2_desc or not cell_5cat or not cell_6cat or not cell_7cat or not cell_8cat:
        break

    codigo.cargar_driver('https://wecare.neoris.net/')

    # click en New Ticket ok
    codigo.d_Xpatch("//td[contains(text(), 'New Ticket')]")
    
    # click en Categoria Principal //Tikcet Infomation Technology
    #codigo.find_element_by_id("btn4").click()
    codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[1]/div/div/div/div/div/div/div[5]/div")

    # Selecione CAT principal
    cat_index, cat_number = codigo.sel_cat(cell_5cat)
    codigo.d_Xpatch(cat_index)

    # Seleccionar sub-categoria
    funciones = {
        5: codigo.sel_FA,
        4: codigo.sel_net,
        3: codigo.sel_LS,
        2: codigo.sel_DC,
        1: codigo.sel_CA
    }

    # Seleccionar sub-categoria
    funciones = {
        5: codigo.sel_FA,
        4: codigo.sel_net,
        3: codigo.sel_LS,
        2: codigo.sel_DC,
        1: codigo.sel_CA
    }

    funcion = funciones.get(cat_number, None)
    if funcion:
        cat_index2, cat_number2 = funcion(cell_6cat)
        codigo.d_Xpatch(cat_index2)
# ok
    # Seleccionar Item Final
    funciones2 = {
        5: codigo.sel_FA_final,
        4: codigo.sel_net_final,
        3: codigo.sel_LS3_final,
        2: codigo.sel_DC_final,
        1: codigo.sel_CA_final
    } 

    funcion2 = funciones2.get(cat_number, None)
    if funcion2:
        cat_index3, cat_number3 = funcion2(cell_7cat)
        codigo.d_Xpatch(cat_index3)

    # click en Change User
    codigo.d_Xpatch("/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[6]/div/div[2]/div/div/table/tbody/tr[1]/td[1]/a")
    
    # PONE EL NOMBRE EN EL INPUT
    codigo.Adorisio(cell_8cat)

    # click en nombre search ok
    codigo.d_Xpatch('/html/body/div[12]/div[2]/div[1]/div/div/div/div/div[2]/div[1]/div/div/div/fieldset/div/div[2]/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/em/button')

    # click en PRIMERA BUSQUEDA del search ok
    codigo.d_Xpatch('/html/body/div[12]/div[2]/div[1]/div/div/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div[2]/div[1]/div/div[1]/div[2]/div/div/table/tbody/tr[1]/td[2]/div')
    # CLICK EN EL SELECT DE LA BUSQUEDA
    codigo.d_Xpatch('/html/body/div[18]/ul/li/a') 

    # Seleccionar Urgency
    #codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[6]/div/div[2]/div/div/table/tbody/tr[2]/td[2]/select')
    
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[6]/div/div[2]/div/div/table/tbody/tr[2]/td[2]/select/option[3]')
    
    # Seleccionar Impact
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[6]/div/div[2]/div/div/table/tbody/tr[2]/td[4]/select/option[3]')

    # Title
    codigo.d_SKcss(cell1_title, 'input#txtProblem')

    # Description
    codigo.d_ID_text(cell2_desc, 'txtDescription')

    # Seleccionar Save
    #codigo.d_Xpatch("//td[contains(text(), 'Save')]")
    codigo.d_Xpatch('/html/body/table/tbody/tr[3]/td/table/tbody/tr/td[4]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/form/table/tbody/tr[4]/td/div[6]/div/div[2]/div/div/table/tbody/tr[9]/td/table/tbody/tr/td[2]/div/table/tbody/tr[2]/td[2]/em/button')
    
    sheet.cell(row=row, column=col3).value = codigo.Guardar_Num_Ticket(
        'lblTicketNumber')
    
    sheet.cell(row=row, column=col9).value = "realizado"
    
    
    wb.save("tickets_creados.xlsx")
    wb.save("crear_tickets.xlsx")
    # Incrementa la fila
    row += 1

    # Cierra el Driver
    codigo.Cerrar_Driver()